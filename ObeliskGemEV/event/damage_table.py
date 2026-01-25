"""
Generate a table showing enemy damage per wave.
"""

import os
from .stats import EnemyStats

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def round_number(number: float, precision: int = 0) -> float:
    """Round a number to specified precision"""
    return round(number, precision)


def calculate_enemy_damage(enemy: EnemyStats, wave: int) -> dict:
    """Calculate enemy damage values for a given wave"""
    # Base damage
    base_dmg = max(1, round_number(enemy.atk + wave * enemy.atk_scaling))
    
    # Crit multiplier
    crit_mult = enemy.crit_dmg + enemy.crit_dmg_scaling * wave
    
    # Crit damage (only if crit_mult > 1)
    if crit_mult > 1:
        crit_dmg = round_number(base_dmg * crit_mult)
    else:
        crit_dmg = base_dmg
    
    # Crit chance
    crit_chance = enemy.crit + wave
    
    return {
        'wave': wave,
        'base_dmg': base_dmg,
        'crit_mult': crit_mult,
        'crit_dmg': crit_dmg,
        'crit_chance': crit_chance
    }


def print_damage_table(max_wave: int = 50):
    """Print a formatted table of enemy damage per wave"""
    enemy = EnemyStats()
    
    print("\n" + "="*80)
    print("GEGNER SCHADEN PRO WAVE (Basiswerte - ohne Upgrades, Prestige 0)")
    print("="*80)
    print(f"{'Wave':<6} {'Base DMG':<10} {'Crit Mult':<12} {'Crit DMG':<10} {'Crit Chance':<12}")
    print("-"*80)
    
    for wave in range(1, max_wave + 1):
        dmg_info = calculate_enemy_damage(enemy, wave)
        crit_chance_str = f"{dmg_info['crit_chance']}%" if dmg_info['crit_chance'] > 0 else "0%"
        crit_mult_str = f"{dmg_info['crit_mult']:.2f}x" if dmg_info['crit_mult'] > 1 else "1.00x"
        base_dmg_str = f"{int(dmg_info['base_dmg'])}"
        crit_dmg_str = f"{int(dmg_info['crit_dmg'])}"
        
        print(f"{dmg_info['wave']:<6} {base_dmg_str:<10} {crit_mult_str:<12} {crit_dmg_str:<10} {crit_chance_str:<12}")
    
    print("="*80)
    print(f"\nBasiswerte:")
    print(f"  Base ATK: {enemy.atk}")
    print(f"  ATK Scaling: +{enemy.atk_scaling} pro Wave")
    print(f"  Base Crit Dmg: {enemy.crit_dmg}x")
    print(f"  Crit Dmg Scaling: +{enemy.crit_dmg_scaling} pro Wave")
    print(f"  Base Crit Chance: {enemy.crit}%")
    print(f"  Crit Chance Scaling: +1% pro Wave")
    print()


def save_damage_table_to_excel(max_wave: int = 50, filename: str = "enemy_damage_table.xlsx"):
    """Save the damage table to an Excel file"""
    if not OPENPYXL_AVAILABLE:
        print("Fehler: openpyxl ist nicht installiert. Bitte installiere es mit: pip install openpyxl")
        return False
    
    enemy = EnemyStats()
    wb = Workbook()
    ws = wb.active
    ws.title = "Gegner Schaden"
    
    # Header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    headers = ["Wave", "Base DMG", "Crit Mult", "Crit DMG", "Crit Chance (%)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for wave in range(1, max_wave + 1):
        dmg_info = calculate_enemy_damage(enemy, wave)
        row = wave + 1
        
        ws.cell(row=row, column=1, value=dmg_info['wave']).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=int(dmg_info['base_dmg'])).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=round(dmg_info['crit_mult'], 2)).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=int(dmg_info['crit_dmg'])).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=dmg_info['crit_chance']).alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    # Add info sheet
    info_ws = wb.create_sheet("Info")
    info_ws['A1'] = "Basiswerte (ohne Upgrades, Prestige 0)"
    info_ws['A1'].font = Font(bold=True, size=14)
    
    info_data = [
        ["Base ATK:", enemy.atk],
        ["ATK Scaling:", f"+{enemy.atk_scaling} pro Wave"],
        ["Base Crit Dmg:", f"{enemy.crit_dmg}x"],
        ["Crit Dmg Scaling:", f"+{enemy.crit_dmg_scaling} pro Wave"],
        ["Base Crit Chance:", f"{enemy.crit}%"],
        ["Crit Chance Scaling:", "+1% pro Wave"],
    ]
    
    for i, (label, value) in enumerate(info_data, 2):
        info_ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        info_ws.cell(row=i, column=2, value=value)
    
    info_ws.column_dimensions['A'].width = 20
    info_ws.column_dimensions['B'].width = 25
    
    # Save file
    filepath = os.path.join(os.path.dirname(__file__), "..", "..", filename)
    wb.save(filepath)
    print(f"\nTabelle gespeichert als: {os.path.abspath(filepath)}")
    return True


if __name__ == "__main__":
    print_damage_table(50)
    print("\nSpeichere als Excel...")
    save_damage_table_to_excel(50)
